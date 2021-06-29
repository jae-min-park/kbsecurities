# -*- coding: utf-8 -*-
"""
Created on Wed Jan 30 13:45:07 2019

@author: infomax
"""
import datetime
import pandas as pd
from pandas import Series, DataFrame
from collections import OrderedDict
import random
import os
import numpy as np

#전체일자 분봉데이터에서 일간data 생성
def candles_to_daily (df=DataFrame):
    #unique date list 생성
    ld = get_date_list(df)
    
    #결과저장 df
    df_daily = DataFrame(columns=['일자','시가','고가',
                                  '저가', '종가'])
    # df_daily에 하루의 결과를 하나의 row로 쌓아나감
    for d in ld:
        df_d = df.loc[df['일자']==d,:]
        df_d_sorted = df_d.sort_values(by=['시간'])
        open_prc = df_d_sorted['시가'].iloc[0]
        close_prc = df_d_sorted['종가'].iloc[-1]
        low_prc = df_d_sorted['저가'].min()
        hi_prc = df_d_sorted['고가'].max()
        d_summary = Series({'일자': d, '시가':open_prc, '고가': hi_prc,
                            '저가': low_prc, '종가': close_prc})
        df_daily = df_daily.append(d_summary, ignore_index=True)
    return df_daily
    

#전체일자분봉데이터와 날짜를 받아서 특정일 분봉데이터생성 및 데이터오류체크
def day_slice (df=DataFrame, d=datetime.date):
    df_day = df.loc[df['일자'] == pd.Timestamp(d),:]
    if len(df_day) == 0:
        print("day_slice says: wrong d!!!!!!!!!")
    else:
        return df_day

#특정일분봉데이터와 시간을 받아서 하나의 캔들데이터 생성 및 데이터오류체크
def t_slice (df_day=DataFrame, t=datetime.time):
    df_t = df_day.loc[df_day['시간']==t,:]
    if len(df_day) == 0:
        print("t_slice says: wrong arg df!!!!!!!!!")
    elif len(df_t) == 0:
        print("t_slice says: wrong t!!!!!!!!!")
    else:
        return df_t
    

#파일명 인자를 받으면 파일명 뒤에 시간을 붙여서 돌려주는 함수
#def f_name_time(filename=str):
#    now= datetime.datetime.now()
#    f_time = f'{now.hour}'+"_"+ f'{now.minute}' + "_" + f'{now.second}'
#    file_name_t = filename + f_time+'.xlsx'
#    return file_name_t

#DataFrame 또는 Series 값을 datetime.time 으로 변환해주는 함수
def to_datetime_time(df=Series):
    df = pd.to_datetime(df)
    df = pd.Series([val.time() for val in df])
    return df

def read_excel_inorder(filename=str): #filename은 full path로
    """엑셀자료 읽어오면서 일자,시간순 정렬"""
    df = pd.read_excel(filename)
    df = df.sort_values(['일자', '시간'], ascending=[True, True])
    df.index = df['일자']
    df.index.name = 'date'
    return df

def read_excel_ust(filename, sheet_name, header, usecols):
    """
     - UST data정리하면서 변경해서 함수 이름이 _ust임.. 그냥 쓰자
     - 정렬 완료했으므로 이후 단계에서 정렬은 필요없음
     - index는 '일자' 컬럼에서 가져온 일자로 하고 index name은 'date'로
    
    """
    dfust = pd.read_excel(filename, sheet_name=sheet_name, 
                          header=header, usecols=usecols)
    dfust = dfust.dropna()
    dfust = dfust.sort_values(['일자', '시간'], ascending=[True, True])
    dfust.index = dfust['일자']
    pxcolname = dfust.columns[-1] #마지막 컬럼이 종가열임
    dfust = dfust.rename(columns = {pxcolname:'종가'})
    dfust.index.name = 'date'
    return dfust

def get_date_list(df=DataFrame):
    """
    데이터를 받아 일자 리스트 리턴
    
    returns: datelist sorted as [-1] being max 
    """
    if '일자' in df.columns:
        ld = sorted(set(list(df['일자']))) #정렬(중복제거(리스트화()))
    else:
        ld_temp = sorted(list(set(df.index.date)))
        ld = [pd.Timestamp(d) for d in ld_temp]
        
    return ld

def jongga(df=DataFrame, date=pd.Timestamp):
    """
    mkt df에서 특정일의 종가 리턴
    """
    if standard_length_of_day(df) == 1:
        return df.loc[date]['종가']
    else:
        return df.loc[date]['종가'][-1]

def juniljongga(ld = list, df=DataFrame, date=pd.Timestamp):
    """mkt df에서 특정일의 전일종가 리턴"""
    junil_date = date_offset(ld, date, -1)
    junil_closeprc = jongga(df, junil_date)
    return junil_closeprc

def get_jongga_dict(df=DataFrame):
    """분봉데이터를 받아 일별종가 dict리턴"""
    ld = get_date_list(df)
    jongga_dict = dict()
    for date in ld: #당일종가 dict 생성
        close_prc = jongga(df, date)
        jongga_dict[date] = close_prc
    return jongga_dict

def get_siga_dict(df=DataFrame):
    """분봉데이터를 받아 일별시가 dict리턴"""
    ld = get_date_list(df)
    siga = dict()
    for d in ld: #당일시가 dict 생성
        df_d = df.loc[df['일자']==d,:]
        df_d_sorted = df_d.sort_values(by=['시간'])
        open_prc = df_d_sorted['시가'].iloc[0]
        siga[d] = open_prc
    return siga

def get_siga_junilbi_dict(df=DataFrame):
    """분봉데이터를 받아 일별 시가전일비 dict리턴"""
    """BC 정의 필요!!!!!!!!!!! 오류발생상태"""
    ld = get_date_list(df)
    ld = ld[1:]
    siga = get_siga_dict(df)
    junil_jongga = get_junil_jongga_dict(df)
    siga_junilbi = dict()
    for d in ld: #당일시가의 전일비 dict 생성
        siga_junilbi[d] = siga[d] - junil_jongga[d]
    return siga_junilbi

def get_junil_jongga_dict(df=DataFrame):
    """분봉데이터를 받아 일별전일종가 dict리턴"""
    ld = get_date_list(df)
    jongga = get_jongga_dict(df)
    junil_jongga = dict()
    for today, prev in zip(ld[1:], ld): #당일종가로부터 전일종가 dict 생성
        junil_jongga[today] = jongga[prev]
    return junil_jongga

def add_junil_jongga_col(df=DataFrame):
    """분봉데이터를 받아 각 행에 전일종가열을 추가한 df리턴"""
    df['전일종가'] = ""    
    junil_jongga = get_junil_jongga_dict(df)
    for d in junil_jongga.keys(): #분봉데이터에 전일종가 data추가
        df.loc[df['일자']==d, '전일종가'] = junil_jongga[d]
    return df

def date_ofweekday(date, weekday):
    """날짜 리스트와 특정 요일을 받아 특정요일의 날짜 리스트를 반환"""
    s_date =[]
    for x in date:
        day_weekday = x.weekday()
        if weekday == day_weekday :
            s_date.append(x)
    return s_date

def date_ofmonth(date, month_num):
    """날짜 리스트와 특정 월을 받아 특정월의 날짜 리스트를 반환"""
    s_date =[]
    for x in date:
        day_month_num = x.month
        if month_num == day_month_num :
            s_date.append(x)
    return s_date

def date_neighbor(ld=list, ref_day=pd.Timestamp, n=int):
    """날짜 리스트와 기준일, n을 받아 
    기준일 이전(n<0) 또는 이후 n개의 날짜 리스트를 반환
    out of range에 대해서는 가능한 날짜들만 리턴 
    예외처리기능필요!!!!!!!!!!!!!!!!!!!"""
    i = ld.index(ref_day)
    if n >0: ldn = ld[i+1:i+n+1]
    else: ldn = ld[i+n:i]
    return ldn

def standard_length_of_day(df=DataFrame):
    """
    Return number of data in a day, assumming last day represents the entire data
    """
# =============================================================================
#     #속도문제로 변경 2020.9.5
#     df = df[pd.Timestamp(2016,8,1):] #이때부터 종가가 15:45으로 30분 늦춰짐
#     s = 0; n = 5
#     for i in range(n):
#         dfi = random.choice(df.index)
#         s = s + len(df[dfi:dfi])
#     return int(round(s/n, 0))
# =============================================================================
    return len(df[df.index[-1]:df.index[-1]])
#    testday = df.index[-1]
#    return len(df[testday:testday])

def date_offset(ld=list, ref_day=pd.Timestamp, n=int):
    """return nth offset day in a date list"""
    i = ld.index(ref_day)
    return ld[i+n]

def date_offset_list(ld_master=list, dates=list, n=int):
    """return nth offset dates list"""
    dates_offset = []
    for d in dates:
        dates_offset.append(date_offset(ld_master, d, n))
    return dates_offset

def read_casedates(casename=str, 
                   mkt="KR",
                   today=None,
                   nth_in_year=None):
    """
    Return list of dates sorted as oldest at the tail
     i.e. dates[0] refers to the latest
    case name과 동일한 sheet name을 사용
    """
    if mkt == "US":
        filename = os.getcwd() + '\CASE_dates.xlsx'
    elif mkt == "KR":
        filename = os.getcwd() + '\dates.xlsx'
    # elif mkt == "SE":
        # filename = os.getcwd() + '\SE_dates.xlsx'
    else: print("Wrong mkt argument")
    
    df = pd.read_excel(filename, sheet_name=casename, header=1, usecols="A")
    df = df.dropna()
    df.columns = ['dates']
    
    df = df.sort_values(by=['dates'], ascending=False)
    df['year'] = [d.year for d in df['dates']]
    df['month'] = [d.month for d in df['dates']]
    df['nth_in_year'] = 0
    for yr in set(df['year']):
        series_nth_in_year = df[df['year']==yr]['dates'].rank(ascending=True) #temporary seires for nth
        for i in series_nth_in_year.index:
            df.loc[i, 'nth_in_year'] = series_nth_in_year[i]
    
    if nth_in_year != None:
        df = df[df['nth_in_year'].isin(nth_in_year)]
    
    
    if today == None:
        dates = list(df.iloc[:,0])
    else: dates = list(df.iloc[:,0]) + [today]
    
    dates.sort(reverse=True)
    
    return dates


#수정필요
def zscore_find(ld=list, df=DataFrame, zsdate=pd.Timestamp, zswindow=20, prc=None, ):
    """return zscore of price (or junil close prc)"""
    window_enddate = date_offset(ld, zsdate, -1)
    window_startdate = date_offset(ld, zsdate, -zswindow)
    df_window = df.loc[(df['일자'] >= window_startdate) & (df['일자'] <= window_enddate),:]
    jong_dict_window = get_jongga_dict(df_window)
    
    windowma = np.mean(list(jong_dict_window.values()))
    windowstd = np.std(list(jong_dict_window.values()))
    
    if prc == None: prc = juniljongga(df, zsdate)
    
    return (prc - windowma) / windowstd

def zscore_find_daily(ld_all, df=DataFrame, ref_date=pd.Timestamp, zswindow=20):
    """
    Return zscore of daily market data
    Zscore of n days upto ref_date-1
    """
    enddate = date_offset(ld_all, ref_date, -1)
    startdate = date_offset(ld_all, ref_date, -zswindow)
    df_window = df[startdate:enddate]
    
    window_avg = np.mean(df_window['종가'])
    window_stdev = np.std(df_window['종가'], ddof=0)
    
#    print(window_avg, "  ",window_stdev)
    
    junil_jongga = df.loc[enddate]['종가']
    
    zs = round((junil_jongga - window_avg) / window_stdev, 2)
    
    return zs
    


class Market:
    """
    2020.1.5 log
    ld, zscore등 market을 class로 해야할 이유는 많음
    pu에서도 Market class를 사용하기 위해서는 df를 호출하는 곳마다 변경이 필요함
    일단 refresh같은 급선무인 기능을 여기에 구현하고, 차차 변경해 가자 
    10분봉으로 변경하면 loadmkt 시간 소요 클 것이므로 구현하긴 해야함
    """
    def __init__(self, df=DataFrame, today=pd.Timestamp, gvntime=None):
        self.dfmkt = df
        self.start_day = df.index[0]
        self.today = today
        self.total_market_dates = get_date_list(df)
        self.junil_jongga_series = Series(get_junil_jongga_dict(df))
        #self.px_last = 현재가

        pass
    
    def refresh(self):
        pass
    
    pass
    
def days_between_dates(day1, day2, ld_all=list):
    """
    Return distance between two dates given entire date list
    ld_all should be sorted, but order doesn't matter
    """
    if (day1 in ld_all) and (day2 in ld_all):
        distance = abs(ld_all.index(day1) - ld_all.index(day2))
        return distance
    else:
        print("wrong date")


def remove_coronavoldays_from_datelist(datelist=list):
    #코로나 변동성 달 제거
    yyyymm = [202003]
    datelist = sorted(set(datelist) - set(select_yyyymm(datelist, yyyymm)), reverse=True)
        
    return datelist

def remove_yyyymm(datelist=list, yyyymm=list):
    """
    datelist를 받아서 관심있는 yyyy-mm삭제
    """
    datelist = sorted(set(datelist) - set(select_yyyymm(datelist, yyyymm)), reverse=True)
        
    return datelist

def select_months(datelist=list, months=list):
    """
    datelist를 받아서 관심있는 달만 남기고 나머지 삭제
    """
    set_mon = set(months)
    datelist = [day for day in datelist if day.month in set_mon]
    datelist.sort(reverse=True)

    return datelist

def select_years(datelist=list, years=list):
    """
    datelist를 받아서 관심있는 year만 남기고 나머지 삭제
    """
    set_yr = set(years)
    datelist = [day for day in datelist if day.year in set_yr]
    datelist.sort(reverse=True)

    return datelist

def select_yyyymm(datelist=list, yyyymm=list):
    """
    datelist를 받아서 관심있는 yyyy-mm만 남기고 나머지 삭제
    """
    set_ym = set(yyyymm)
    datelist = [day for day in datelist if (day.year*100+day.month) in set_ym]
    datelist.sort(reverse=True)
    
    return datelist

def sort_zs_similarity(dfmkt=DataFrame, datelist=list, zscutoff=float):
    """
    datelist를 Zscore 유사성 기준으로 정렬
    
    parameter
    datelist  datelist sorted by timely order
    zswin : Zscore window
    """
    ZSWIN1 = 20
    # ZSWIN2 = 20
#    ZSWIN3 = 60
    ld_all = get_date_list(dfmkt)
    df = pd.DataFrame(index = datelist, columns=['zs1', 'zs2', 'zs3', 'zs_diff', 'sim_rank'])
    zs1_of_latest = zscore_find_daily(ld_all, dfmkt, datelist[0], ZSWIN1)
    # zs2_of_latest = zscore_find_daily(ld_all, dfmkt, datelist[0], ZSWIN2)
#    zs3_of_latest = zscore_find_daily(ld_all, dfmkt, datelist[0], ZSWIN3)
    
    for dti in df.index:
        df.loc[dti]['zs1'] = zscore_find_daily(ld_all, dfmkt, dti, ZSWIN1)
#        df.loc[dti]['zs2'] = zscore_find_daily(ld_all, dfmkt, dti, ZSWIN2)
#        df.loc[dti]['zs3'] = zscore_find_daily(ld_all, dfmkt, dti, ZSWIN3)

        df.loc[dti]['zs_diff'] = (zs1_of_latest - df.loc[dti]['zs1'])**2
#        + (zs2_of_latest - df.loc[dti]['zs2'])**2
#        + (zs3_of_latest - df.loc[dti]['zs3'])**2
        
    df['sim_rank'] = df['zs_diff'].rank(ascending=True)
    df = df.sort_values(by=['sim_rank'])
        
    datelist_sorted = list(df.index)[:int(len(list(df.index))*zscutoff)]
    
    return datelist_sorted


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
            
    
def common_elements(list1, list2):
    """
    return common elements in two lists
    """
    return sorted(list(set(list1).intersection(list2)), reverse=True)

def dates_from_filer(filtered_df, corona_exemption = "Y"):
    dates = sorted(list(filtered_df.index), reverse=True)
    if corona_exemption == "Y":
        dates = remove_coronavoldays_from_datelist(dates)
    return dates


def days_to_event(ld_all, today, eventdates):
    """
    Return number of days remaining until event date.
    
    Parameters
    -----------
    ld_all : list of total market dates 
    eventdates : list of dates of interest  
    """
    
    upcoming_event_date = pd.Timestamp(2020,2,27)
    remaining_days = days_between_dates(today, upcoming_event_date, ld_all)
    
    return remaining_days

def remove_neighbor_dates(ld_all=list, datelist=list, closer_than=5):
    """
    Remove neighboring dates from a list of dates [datelist].
    Remains only the latest date among close dates.
    
    Parameters:
      ld_all : list of all dates. Used to calculate distance between dates.
      datelist : list of dates containing neighboring dates.
      closer_than : date removal threshold.

    Returns: datelist
    2020.11.03 
    """
    datelist = sorted(datelist, reverse=True)
    
    for ref_day in datelist:
        for test_day in [d for d in datelist if d < ref_day]:
            distance = days_between_dates(ref_day, test_day, ld_all)
            if distance < 5: datelist.remove(test_day)    
    return datelist

def dt(date=int):
    date_string = str(date)
    return pd.Timestamp(int(date_string[0:4]),
                        int(date_string[4:6]),
                        int(date_string[6:8]))   


def last_mktday(dfmkt, refday=None):
    """
    Returns last businessday before refday.
    Parameters:
        ld_all : mkt datelist master 
        refday : if None, today-1 will be assigned
    """
    yday = datetime.date.today() - datetime.timedelta(days=1)
    
    if refday == None: 
        refday = yday
    else: 
        refday = refday - datetime.timedelta(days=1)
        
    return dfmkt[:refday].index[-1]

def next_weekday(d, weekday):
    days_ahead = weekday - d.weekday()
    if days_ahead < 0: # Target day already happened this week
        days_ahead += 7
    return d + datetime.timedelta(days_ahead)

def prev_weekday(d, weekday):
    days_passed = d.weekday() - weekday
    if days_passed < 0: # Target day has not yet come
        days_passed += 7
    return d - datetime.timedelta(days_passed)


def align_weekday_early(datelist, target_weekday=0):
    """
    Adjust dates in datelist to be the latest weekday.
    ex) weekday = 0, Mon~Fri -> Mon in that week
    Parameters:
        ld_all : mkt datelist master 
        datelist : to-be-aligned datelist
    Returns:
        date list aligned to same weekday, Monday by default
    """
    return [prev_weekday(d, target_weekday) for d in datelist ]
        
def align_to_near_monday(datelist):
    """
    Adjust dates in datelist to be the same weekday nearest
    ex) target_weekday=0: Tue -> last Mon, Fri -> next Mon
    Needed to align US auction days to Monday as US Treasury auction date varies
    Parameters:
        datelist : to-be-aligned datelist
    Returns:
        date list aligned to same weekday, Monday by default
    """
    datelist_aligned = []
    for d in datelist:
        if d.weekday() in [0,1,2]:
            datelist_aligned.append(prev_weekday(d, 0))
        elif d.weekday() in [3,4]:
            datelist_aligned.append(next_weekday(d, 0))
        else: raise Exception("Weekend days found in datelist")
        
    return datelist_aligned
            
    
    
    
    
    
    



